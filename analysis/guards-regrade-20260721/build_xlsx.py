#!/usr/bin/env python3
"""Build the hermia corpus/grader/GUARDS provenance tracker workbook."""
from __future__ import annotations

import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

SP = Path("/private/tmp/claude-501/-Users-scottbly-Git-hermia/8227e715-a5ca-4794-8900-ddb7d0290fe8/scratchpad")
OUT = Path("/Users/scottbly/Git/hermia/docs/hermia-corpus-provenance-tracker.xlsx")

D = json.loads((SP / "reconstruction.json").read_text())

FONT = "Arial"
HDR_FILL = PatternFill("solid", fgColor="1F3864")
HDR_FONT = Font(name=FONT, bold=True, color="FFFFFF", size=10)
TITLE_FONT = Font(name=FONT, bold=True, size=14, color="1F3864")
SUB_FONT = Font(name=FONT, italic=True, size=10, color="595959")
BODY = Font(name=FONT, size=10)
BOLD = Font(name=FONT, size=10, bold=True)
INPUT_FONT = Font(name=FONT, size=10, color="0000FF")

GUARDS_FILL = PatternFill("solid", fgColor="E2EFDA")   # green  = GUARDS era
PRE_FILL = PatternFill("solid", fgColor="FCE4D6")      # orange = pre-GUARDS
WARN_FILL = PatternFill("solid", fgColor="FFF2CC")     # yellow = attention
THIN = Side(style="thin", color="BFBFBF")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

wb = Workbook()


def header_row(ws, row, headers, widths=None):
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=i, value=h)
        c.fill = HDR_FILL
        c.font = HDR_FONT
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BOX
    if widths:
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = ws.cell(row=row + 1, column=1)


def put(ws, r, c, v, font=BODY, fill=None, fmt=None, align=None):
    cell = ws.cell(row=r, column=c, value=v)
    cell.font = font
    cell.border = BOX
    if fill:
        cell.fill = fill
    if fmt:
        cell.number_format = fmt
    if align:
        cell.alignment = Alignment(horizontal=align, vertical="top", wrap_text=True)
    else:
        cell.alignment = Alignment(vertical="top", wrap_text=True)
    return cell


# ══════════════════════════════════════════════════════════ 1. README
ws = wb.active
ws.title = "README"
ws.column_dimensions["A"].width = 3
ws.column_dimensions["B"].width = 34
ws.column_dimensions["C"].width = 104

r = 2
ws.cell(row=r, column=2, value="Hermia — Corpus / Grader / GUARDS Provenance Tracker").font = TITLE_FONT
r += 1
ws.cell(row=r, column=2, value="Reconstructed 2026-07-21 from git history + evidence preserved inside the run files.").font = SUB_FONT
r += 2

blocks = [
    ("Why this exists",
     "Hermia's version stamps do NOT track when the tests, graders, or GUARDS actually changed. "
     "The hermia_version field did not exist before ~2026-06-12, and the July runs carry a known "
     "mis-stamp (0.1.0 on 0.1.3 source). Reconstructing this from memory each time was producing "
     "wrong conclusions. Git history was maintained properly, so the truth is recoverable — this "
     "workbook is that reconstruction, so it does not have to be redone."),
    ("The key correction",
     "You cannot select the pre-GUARDS era by version number. By the time anything was stamped "
     "0.1.3 (first appears 2026-06-12), GUARDS was ALREADY in the corpus (landed 2026-06-09). "
     "Every genuinely pre-GUARDS run carries NO version stamp at all. Select by corpus era or "
     "date, never by version_stamp."),
    ("How corpus version is identified",
     "Every result row stores raw_system — the system prompt actually sent to the model. Hashing "
     "those and comparing against the corpus as it existed at each git commit identifies the "
     "corpus version from EVIDENCE, not from a date guess. 73 of 76 runs matched at 100%."),
    ("Confidence: what is proven vs inferred",
     "CORPUS provenance = VERIFIED where the prompt hashes match (marked per run). "
     "GRADER provenance = INFERRED — derived by date ordering against schemas.py history, NOT "
     "fingerprinted. A run's grader is the most recent schemas.py commit predating it, which "
     "assumes the executing machine had that commit checked out. Treat grader attribution as a "
     "strong assumption, not proof."),
    ("Committed \u2260 in use",
     "The Corpus Versions sheet dates each version by when it was COMMITTED, which is not when it "
     "was in use. The 28-test corpus was running from 2026-05-18 but was not committed until "
     "2026-06-03 \u2014 roughly two weeks of runs on content that git dates later. Use the "
     "hash-matched corpus era (per run), never the commit date, to decide what a run actually used."),
    ("Stale checkouts are real \u2014 check the flag",
     "Two runs (2026-06-23 and 2026-06-24) POSTDATE the GUARDS commit but ran a pre-GUARDS corpus "
     "from a stale checkout \u2014 prompts match C12 at 28/28. Classifying runs by date put "
     "pre-GUARDS prompts in the post-GUARDS group. The Runs sheet flags these as STALE. Always "
     "bucket by hash-verified era."),
    ("Known limitation",
     "The corpus GREW from 8 tests (0 security) to 30 tests (15 security) across this period. Any "
     "pre/post comparison must intersect on test_id, or it measures corpus growth rather than "
     "prompt quality."),
]
for title, body in blocks:
    ws.cell(row=r, column=2, value=title).font = BOLD
    c = ws.cell(row=r, column=3, value=body)
    c.font = BODY
    c.alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[r].height = 15 * (len(body) // 100 + 1) + 12
    r += 1

r += 1
ws.cell(row=r, column=2, value="Sheet guide").font = BOLD
r += 1
for name, desc in [
    ("Timeline", "Every corpus / grader / GUARDS commit in one chronological list."),
    ("Corpus Versions", "The 18 distinct corpus states (C01–C18), test counts, GUARDS state."),
    ("Grader Versions", "schemas.py commit history — the grading logic lineage."),
    ("Runs", "All 76 result files mapped to their reconstructed corpus + grader version."),
    ("GUARDS Evidence", "The 2026-07-21 re-grade result: measured GUARDS effect and its limits."),
]:
    ws.cell(row=r, column=2, value=name).font = Font(name=FONT, size=10, bold=True, color="1F3864")
    ws.cell(row=r, column=3, value=desc).font = BODY
    r += 1

r += 1
ws.cell(row=r, column=2, value="Live counts (formulas — update if rows are added)").font = BOLD
r += 1
counts = [
    ("Result files tracked", "=COUNTA(Runs!A3:A200)"),
    ("Runs on the pre-GUARDS corpus", '=COUNTIF(Runs!U3:U200,"pre-GUARDS")'),
    ("Runs on the GUARDS corpus", '=COUNTIF(Runs!U3:U200,"GUARDS")'),
    ("Runs on a STALE pre-GUARDS checkout", '=COUNTIF(Runs!F3:F200,"STALE")'),
    ("Runs with VERIFIED corpus provenance", '=COUNTIF(Runs!V3:V200,"VERIFIED*")'),
    ("Distinct corpus versions", "=COUNTA('Corpus Versions'!A3:A100)"),
    ("Total evaluation rows", "=SUM(Runs!D3:D200)"),
]
for label, formula in counts:
    ws.cell(row=r, column=2, value=label).font = BODY
    c = ws.cell(row=r, column=3, value=formula)
    c.font = BOLD
    c.alignment = Alignment(horizontal="left")
    r += 1

r += 1
ws.cell(row=r, column=2, value="GUARDS landing").font = BOLD
ws.cell(row=r, column=3,
        value=f"{D['guards_landing_date'][:10]}  •  commit 067f141  •  "
              '"normalize(corpus): GUARDS 6/6 for all 30 tests + adversarial framing for '
              'multi-step and numeric" (PR #105).  Follow-up 0bfe26a (PR #127) 2026-06-26 '
              "added the GUARDS-R refusal path.").font = BODY
ws.row_dimensions[r].height = 44

# ══════════════════════════════════════════════════════════ 2. TIMELINE
ws = wb.create_sheet("Timeline")
ws.cell(row=1, column=1, value="Chronological history — corpus, grader, and GUARDS").font = TITLE_FONT
header_row(ws, 2,
           ["Date", "Type", "Commit", "Corpus era", "GUARDS state", "Subject"],
           [12, 15, 11, 11, 13, 96])

events = []
for v in D["corpus_versions"]:
    events.append((v["date"][:10], "CORPUS", v["short"], v["corpus_era"], v["guards_state"], v["subject"]))
for c in D["grader_commits"]:
    events.append((c["date"][:10], "GRADER", c["sha"][:7], "", "", c["subject"]))
for c in D["guards_commits"]:
    events.append((c["date"][:10], "GUARDS DOC", c["sha"][:7], "", "", c["subject"]))
for c in D["normalize_commits"]:
    events.append((c["date"][:10], "NORMALIZE", c["sha"][:7], "", "", c["subject"]))
events.sort(key=lambda e: (e[0], e[1]))

r = 3
for date, typ, sha, era, gs, subj in events:
    fill = None
    if typ == "CORPUS":
        fill = GUARDS_FILL if gs == "GUARDS" else PRE_FILL
    elif typ == "GUARDS DOC":
        fill = WARN_FILL
    put(ws, r, 1, date, fill=fill)
    put(ws, r, 2, typ, font=BOLD, fill=fill)
    put(ws, r, 3, sha, fill=fill)
    put(ws, r, 4, era, fill=fill, align="center")
    put(ws, r, 5, gs, fill=fill, align="center")
    put(ws, r, 6, subj, fill=fill)
    r += 1

# ══════════════════════════════════════════════════════════ 3. CORPUS VERSIONS
ws = wb.create_sheet("Corpus Versions")
ws.cell(row=1, column=1, value="Distinct corpus states — each is a version of the test set").font = TITLE_FONT
header_row(ws, 2,
           ["Era", "Date", "Commit", "Tests", "Security tests", "GUARDS state",
            "Path at commit", "Corpus file SHA-256", "Subject"],
           [7, 12, 11, 8, 14, 13, 34, 50, 78])
r = 3
for v in D["corpus_versions"]:
    fill = GUARDS_FILL if v["guards_state"] == "GUARDS" else PRE_FILL
    put(ws, r, 1, v["corpus_era"], font=BOLD, fill=fill, align="center")
    put(ws, r, 2, v["date"][:10], fill=fill)
    put(ws, r, 3, v["short"], fill=fill)
    put(ws, r, 4, v["n_tests"], fill=fill, align="center")
    put(ws, r, 5, v["security_tests"], fill=fill, align="center")
    put(ws, r, 6, v["guards_state"], font=BOLD, fill=fill, align="center")
    put(ws, r, 7, v["path_at_commit"], fill=fill)
    put(ws, r, 8, v["corpus_file_sha256"], fill=fill)
    put(ws, r, 9, v["subject"], fill=fill)
    r += 1

# NB: keep column A clear below the data — README's COUNTA(A3:A100) counts this column.
note_r = r + 1
ws.cell(row=note_r, column=2,
        value="Note: the corpus grew 8 → 30 tests and 0 → 15 security tests over this period. "
              "Any pre/post comparison MUST intersect on test_id or it measures corpus growth.").font = SUB_FONT

# ══════════════════════════════════════════════════════════ 4. GRADER VERSIONS
ws = wb.create_sheet("Grader Versions")
ws.cell(row=1, column=1, value="Grading logic lineage — src/hermia/schemas.py").font = TITLE_FONT
ws.cell(row=2, column=1,
        value="Grader attribution for a run is INFERRED by date ordering, not fingerprinted. "
              "See README.").font = SUB_FONT
header_row(ws, 3, ["#", "Date", "Commit", "Author", "Subject"], [5, 12, 11, 18, 100])
r = 4
for i, c in enumerate(sorted(D["grader_commits"], key=lambda x: x["date"]), 1):
    put(ws, r, 1, i, align="center")
    put(ws, r, 2, c["date"][:10])
    put(ws, r, 3, c["sha"][:7])
    put(ws, r, 4, c["author"])
    put(ws, r, 5, c["subject"])
    r += 1

# ══════════════════════════════════════════════════════════ 5. RUNS
ws = wb.create_sheet("Runs")
ws.cell(row=1, column=1, value="Every result file, mapped to its reconstructed corpus and grader version").font = TITLE_FONT
headers = ["Run date", "Run time", "File", "Rows", "Tests run", "Stale checkout?", "Infra rows", "Gradeable rows",
           "Recorded pass %", "Version stamp", "Corpus stamp", "git_sha stamp",
           "Models", "Hosts", "Top models", "Dimensions",
           "Corpus era", "Corpus commit", "Match", "Match score", "GUARDS state",
           "Corpus provenance", "Grader commit", "Grader date", "Grader provenance"]
widths = [11, 10, 34, 8, 10, 14, 9, 10, 12, 11, 12, 11, 8, 7, 40, 44, 10, 12, 9, 10, 13, 20, 12, 11, 20]
header_row(ws, 2, headers, widths)

r = 3
for run in D["runs"]:
    gs = run.get("guards_state")
    fill = GUARDS_FILL if gs == "GUARDS" else (PRE_FILL if gs == "pre-GUARDS" else WARN_FILL)
    prov = run.get("corpus_provenance") or ""
    vals = [
        run["run_date"], run["run_time"], run["file"], run["rows"],
        run["distinct_tests_run"], "STALE" if run.get("stale_checkout") else "",
        run["infra_rows"],
        run["gradeable_rows"], run["recorded_pass_pct"], run["version_stamp"],
        run["corpus_sha256_stamp"], run["git_sha_stamp"], run["n_models"], run["n_hosts"],
        run["top_models"], run["dimensions"], run["corpus_era"], run["corpus_commit"],
        run["corpus_matched_tests"], run["corpus_match_score"], gs, prov,
        run["grader_commit"], run["grader_commit_date"], run["grader_provenance"],
    ]
    for i, v in enumerate(vals, 1):
        f = BODY
        if i in (6, 17, 21):
            f = BOLD
        cell = put(ws, r, i, v, font=f, fill=fill)
        if i == 6 and v:
            cell.fill = PatternFill("solid", fgColor="F8CBAD")
        if i == 9 and v is not None:
            cell.number_format = "0.00"
        if i == 20 and v is not None:
            cell.number_format = "0.0%"
        if i == 22 and prov and not prov.startswith("VERIFIED"):
            cell.fill = WARN_FILL
            cell.font = BOLD
    r += 1

ws.auto_filter.ref = f"A2:Y{r-1}"

# ══════════════════════════════════════════════════════════ 6. GUARDS EVIDENCE
ws = wb.create_sheet("GUARDS Evidence")
ws.column_dimensions["A"].width = 3
ws.column_dimensions["B"].width = 40
ws.column_dimensions["C"].width = 18
ws.column_dimensions["D"].width = 18
ws.column_dimensions["E"].width = 62

r = 2
ws.cell(row=r, column=2, value="Measured GUARDS effect — 2026-07-21 re-grade").font = TITLE_FONT
r += 1
ws.cell(row=r, column=2,
        value="Full method and limits: docs/provenance/2026-07-21-guards-effect-regrade-analysis.md").font = SUB_FONT
r += 2

ws.cell(row=r, column=2, value="Method").font = BOLD
ws.cell(row=r, column=5,
        value="All 76 historical runs re-graded with the CURRENT grader using stored raw_response. "
              "A = pre-GUARDS/old grader, B = pre-GUARDS/NEW grader, C = post-GUARDS/NEW grader. "
              "C−B matched on (model, test_id) present both sides.").font = BODY
ws.row_dimensions[r].height = 46
r += 2

header_row(ws, r, ["", "Finding", "Value", "95% CI / detail", "Why it matters"], [3, 40, 18, 18, 62])
r += 1

findings = [
    ("Grader effect on pre-GUARDS data", "-1.86 pp", "173 P\u2192F, 0 F\u2192P",
     "The current grader is MONOTONICALLY STRICTER. It never rescues a failure, so it cannot "
     "manufacture an improvement \u2014 any gain survives a tougher yardstick. Load-bearing fact."),
    ("GUARDS effect \u2014 overall (matched)", "+4.97 pp", "[+3.85, +6.08], p=3.5e-16",
     "Same grader, same models, same tests. n = 9,085 pre / 4,316 post. 811 matched pairs."),
    ("GUARDS effect \u2014 security dimension", "+7.57 pp", "[+6.28, +8.85], p=3.5e-23",
     "The dimension GUARDS targets. n = 4,926 / 2,300."),
    ("Routing dimension", "+15.95 pp", "36.6% \u2192 52.6%",
     "Largest gain; corroborates that the high-fail-rate tests improved most."),
    ("Constraint dimension", "+6.70 pp", "69.2% \u2192 75.9%", "Also a GUARDS-targeted dimension."),
    ("Reasoning dimension", "-4.59 pp", "98.3% \u2192 93.8%",
     "TRADEOFF: capability dimensions drift DOWN. Consistent with defensive instruction "
     "consuming attention budget."),
    ("Memory dimension", "-2.01 pp", "93.5% \u2192 91.5%", "Same tradeoff pattern."),
    ("Domain dimension", "-1.58 pp", "100.0% \u2192 98.4%", "Same tradeoff pattern."),
    ("Tool-use dimension", "-0.70 pp", "98.3% \u2192 97.6%", "Essentially flat."),
    ("Model breadth (security)", "19 of 23 improved", "2 down, 2 flat",
     "Effect is broad, not driven by outliers."),
    ("Prompt-change proof", "28 of 28 changed", "0 identical",
     "System prompts hashed from the run data itself \u2014 corpus change proven without git."),
    ("Classification basis", "hash-verified era", "NOT run date",
     "v2 correction: two runs dated after the GUARDS commit ran a stale pre-GUARDS checkout. "
     "Date-based bucketing understated the effect (security was +7.50 before the fix)."),
]
for name, val, detail, why in findings:
    fill = GUARDS_FILL if not val.startswith("-") else WARN_FILL
    put(ws, r, 2, name, font=BOLD, fill=fill)
    put(ws, r, 3, val, font=BOLD, fill=fill, align="center")
    put(ws, r, 4, detail, fill=fill, align="center")
    put(ws, r, 5, why, fill=fill)
    r += 1

r += 1
ws.cell(row=r, column=2, value="DO NOT CLAIM").font = Font(name=FONT, bold=True, size=11, color="C00000")
r += 1
for bad in [
    'Anything sourced to the "v0.1.3 → v0.2.0" window — corpus hash is identical across it.',
    '"GUARDS caused X" — commit 067f141 bundled adversarial-framing changes with GUARDS.',
    "A single headline number without the tradeoff finding — the tradeoff is the honest shape.",
    "That this replaces the ablation. It does not. It motivates it (hermia-dhqv, hermia-5wc.3).",
]:
    ws.cell(row=r, column=2, value="•").font = BODY
    c = ws.cell(row=r, column=3, value=bad)
    c.font = BODY
    c.alignment = Alignment(wrap_text=True)
    r += 1

r += 1
ws.cell(row=r, column=2, value="Uncontrolled variables").font = BOLD
r += 1
for lim in [
    "Hosts/backends: 11 pre, 15 post, only 8 shared. Hardware rides along with time.",
    "Ollama version skew across May→July is uncontrolled and known to be substantial.",
    "Unequal n (8,729 vs 4,673) and unequal repeat counts.",
    "No per-host stratification. Matched analysis controls model and test, not environment.",
    "Infra-failure exclusion rates differ between periods.",
]:
    ws.cell(row=r, column=2, value="•").font = BODY
    c = ws.cell(row=r, column=3, value=lim)
    c.font = BODY
    c.alignment = Alignment(wrap_text=True)
    r += 1

for sheet in wb.worksheets:
    for row in sheet.iter_rows():
        for cell in row:
            if cell.font and cell.font.name != FONT:
                cell.font = Font(name=FONT, size=cell.font.size or 10,
                                 bold=cell.font.bold, italic=cell.font.italic,
                                 color=cell.font.color)

wb.save(OUT)
print(f"wrote {OUT}")
print(f"  sheets: {[s.title for s in wb.worksheets]}")
print(f"  runs: {len(D['runs'])}  corpus versions: {len(D['corpus_versions'])}  "
      f"grader commits: {len(D['grader_commits'])}")
